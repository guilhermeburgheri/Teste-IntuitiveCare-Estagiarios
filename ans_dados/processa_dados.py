import re 
import csv
import zipfile
from pathlib import Path
from openpyxl import load_workbook


CHAVES = ("evento", "sinistro")
SAIDAS_ACEITAS = (".csv", ".txt", ".xlsx", ".xlsm")


def extrair_zips(zips_dir: Path, destino: Path) -> None:
    destino.mkdir(parents=True, exist_ok=True)

    for zip_path in zips_dir.glob("*.zip"):
        pasta = destino / zip_path.stem
        pasta.mkdir(exist_ok=True)

        with zipfile.ZipFile(zip_path, "r") as z:
            z.extractall(pasta)


def slug(s: str) -> str:
    s = (s or "").strip().lower().lstrip("\ufeff").strip('"').strip("'")
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def tem_evento_sinistro(texto: str) -> bool:
    t = (texto or "").lower()
    return any(k in t for k in CHAVES)


def parse_valor(v) -> float | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def linhas_csv_ou_txt(path: Path):
    for delim in (";", ",", "\t"):
        try:
            with path.open("r", encoding="utf-8", errors="ignore", newline="") as f:
                r = csv.DictReader(f, delimiter=delim)
                if r.fieldnames and len(r.fieldnames) > 1:
                    for row in r:
                        yield row
                    return
        except Exception:
            pass


def linhas_xlsx(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            continue

        cols = [str(c).strip() if c is not None else "" for c in header]

        for row in it:
            d = {}
            for i, v in enumerate(row):
                if i >= len(cols):
                    break
                d[cols[i]] = "" if v is None else str(v)
            yield d


def filtrar_eventos_sinistros(base_dir: Path):
    desc_keys = ("descricao", "ds_conta", "descricao_conta", "conta")
    val_keys = ("vl_saldo_final", "valor", "vl", "valor_despesas", "vl_despesas")

    for arq in base_dir.rglob("*"):
        if not arq.is_file() or arq.suffix.lower() not in SAIDAS_ACEITAS:
            continue

        linhas = linhas_csv_ou_txt(arq) if arq.suffix.lower() in (".csv", ".txt") else linhas_xlsx(arq)

        for row in linhas:
            norm = {slug(str(k)): v for k, v in row.items()}

            descricao = ""
            for k in desc_keys:
                if norm.get(k):
                    descricao = str(norm.get(k))
                    break
            if not tem_evento_sinistro(descricao):
                continue

            valor_raw = None
            for k in val_keys:
                if norm.get(k) not in (None, ""):
                    valor_raw = norm.get(k)
                    break

            valor = parse_valor(valor_raw)
            if valor is None:
                continue

            yield {"arquivo": str(arq), "descricao": descricao, "valor": valor}


def contar_arquivos_com_eventos(base_dir: Path) -> int:
    arquivos = set()
    for item in filtrar_eventos_sinistros(base_dir):
        arquivos.add(item["arquivo"])
    return len(arquivos)
